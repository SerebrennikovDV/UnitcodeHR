"""Контроллеры HR-аналитики."""
import io

from django.contrib.auth.decorators import login_required
from django.http import FileResponse, JsonResponse
from django.shortcuts import render

from . import services


@login_required
def dashboard(request):
    """Сводный дашборд HR-аналитики."""
    return render(request, 'analytics/dashboard.html', {
        'time_to_hire': services.time_to_hire(),
        'source_of_hire': services.source_of_hire(),
        'funnel': services.funnel_conversions(),
        'cost_per_hire': services.cost_per_hire(),
        'probation': services.probation_pass_rate(),
        'breadcrumbs': [('Дашборд', '/dashboard/'), ('HR-аналитика', None)],
    })


@login_required
def api_funnel(request):
    """JSON-эндпойнт для Chart.js."""
    return JsonResponse({'funnel': services.funnel_conversions()})


@login_required
def export_xlsx(request):
    """Экспорт аналитических метрик в .xlsx."""
    from openpyxl import Workbook
    wb = Workbook()

    ws_main = wb.active
    ws_main.title = 'Сводка'
    ws_main['A1'] = 'Показатель'
    ws_main['B1'] = 'Значение'
    tth = services.time_to_hire()
    ws_main.append(['Среднее time-to-hire (дни)', tth['avg_days_to_hire']])
    ws_main.append(['Закрыто вакансий за период', tth['closed_count']])
    cph = services.cost_per_hire()
    ws_main.append(['Cost-per-hire, ₽', cph['cost_per_hire']])
    pp = services.probation_pass_rate()
    ws_main.append(['Прохождение испытательного, %', pp['rate']])

    ws_sources = wb.create_sheet('Источники найма')
    ws_sources.append(['Источник', 'Кандидаты', 'Наняты', 'Конверсия, %', 'Стоимость / мес.'])
    for r in services.source_of_hire():
        ws_sources.append([r['source'], r['total_candidates'], r['hired'],
                            r['conversion_pct'], r['cost_per_month']])

    ws_funnel = wb.create_sheet('Воронка')
    ws_funnel.append(['Этап', 'Достигнуто откликов', 'Конверсия от старта, %'])
    for r in services.funnel_conversions():
        ws_funnel.append([r['stage'], r['reached'], r.get('conversion_pct', 0)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True,
                        filename='UnitcodeHR_analytics.xlsx',
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
